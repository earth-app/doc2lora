"""Tests for doc2lora parsers."""

import json
import tarfile
import tempfile
import unittest
import zipfile
from pathlib import Path

from doc2lora.parsers import DocumentParser


class TestDocumentParser(unittest.TestCase):
    """Test cases for DocumentParser."""

    def setUp(self):
        """Set up test fixtures."""
        self.parser = DocumentParser()
        self.temp_dir = tempfile.mkdtemp()
        self.temp_path = Path(self.temp_dir)

    def tearDown(self):
        """Clean up test fixtures."""
        import shutil

        shutil.rmtree(self.temp_dir)

    def test_parse_markdown(self):
        """Test parsing markdown files."""
        content = "# Test\n\nThis is a test markdown file."
        md_file = self.temp_path / "test.md"
        md_file.write_text(content)

        result = self.parser.parse_file(md_file)

        self.assertIsNotNone(result)
        self.assertEqual(result["content"], content)
        self.assertEqual(result["extension"], ".md")

    def test_parse_text(self):
        """Test parsing text files."""
        content = "This is a test text file."
        txt_file = self.temp_path / "test.txt"
        txt_file.write_text(content)

        result = self.parser.parse_file(txt_file)

        self.assertIsNotNone(result)
        self.assertEqual(result["content"], content)
        self.assertEqual(result["extension"], ".txt")

    def test_parse_json(self):
        """Test parsing JSON files."""
        data = {"test": "data", "number": 42}
        json_file = self.temp_path / "test.json"
        json_file.write_text(json.dumps(data))

        result = self.parser.parse_file(json_file)

        self.assertIsNotNone(result)
        self.assertIn('"test": "data"', result["content"])
        self.assertEqual(result["extension"], ".json")

    def test_parse_csv(self):
        """Test parsing CSV files."""
        content = "name,age\nJohn,30\nJane,25"
        csv_file = self.temp_path / "test.csv"
        csv_file.write_text(content)

        result = self.parser.parse_file(csv_file)

        self.assertIsNotNone(result)
        self.assertIn("name, age", result["content"])
        self.assertEqual(result["extension"], ".csv")

    def test_parse_directory(self):
        """Test parsing entire directories."""
        # Create test files
        (self.temp_path / "test1.md").write_text("# Test 1")
        (self.temp_path / "test2.txt").write_text("Test 2 content")
        (self.temp_path / "test3.json").write_text('{"test": 3}')

        # Create subdirectory with file
        subdir = self.temp_path / "subdir"
        subdir.mkdir()
        (subdir / "test4.md").write_text("# Test 4")

        results = self.parser.parse_directory(str(self.temp_path))

        self.assertEqual(len(results), 4)
        filenames = [r["filename"] for r in results]
        self.assertIn("test1.md", filenames)
        self.assertIn("test2.txt", filenames)
        self.assertIn("test3.json", filenames)
        self.assertIn("test4.md", filenames)

    def test_subdirectory_labeling(self):
        """Test that subdirectories are used for document labeling."""
        # Create directory structure
        legal_dir = self.temp_path / "legal"
        tech_dir = self.temp_path / "technical"
        marketing_dir = self.temp_path / "marketing" / "campaigns"

        legal_dir.mkdir()
        tech_dir.mkdir()
        marketing_dir.mkdir(parents=True)

        # Create test files in subdirectories
        (legal_dir / "contract.txt").write_text("Legal contract content")
        (tech_dir / "spec.md").write_text("# Technical specification")
        (marketing_dir / "campaign.txt").write_text("Marketing campaign content")
        (self.temp_path / "root_file.txt").write_text("Root level content")

        results = self.parser.parse_directory(str(self.temp_path))

        # Find results by filename
        contract_result = next(r for r in results if r["filename"] == "contract.txt")
        spec_result = next(r for r in results if r["filename"] == "spec.md")
        campaign_result = next(r for r in results if r["filename"] == "campaign.txt")
        root_result = next(r for r in results if r["filename"] == "root_file.txt")

        # Check labels (now combining directory and filename)
        self.assertEqual(contract_result["label"], "legal_contract")
        self.assertEqual(contract_result["category_path"], "legal")

        self.assertEqual(spec_result["label"], "technical_spec")
        self.assertEqual(spec_result["category_path"], "technical")

        self.assertEqual(campaign_result["label"], "campaigns_campaign")
        self.assertEqual(campaign_result["category_path"], "marketing/campaigns")

        self.assertEqual(root_result["label"], "root_root_file")
        self.assertEqual(root_result["category_path"], "")

    def test_unsupported_file_type(self):
        """Test handling of unsupported file types."""
        unsupported_file = self.temp_path / "test.xyz"
        unsupported_file.write_text("This should not be parsed")

        result = self.parser.parse_file(unsupported_file)

        self.assertIsNone(result)

    def test_parse_zip_archive(self):
        """Test parsing ZIP archives containing supported documents."""
        # Create a ZIP file with test documents
        zip_file = self.temp_path / "test_archive.zip"

        with zipfile.ZipFile(zip_file, "w") as zip_ref:
            # Add a markdown file
            zip_ref.writestr(
                "docs/readme.md", "# Test Documentation\n\nThis is a test."
            )
            # Add a text file
            zip_ref.writestr("notes.txt", "Important notes here.")
            # Add a JSON file
            zip_ref.writestr("config.json", '{"version": "1.0", "name": "test"}')
            # Add an unsupported file
            zip_ref.writestr("image.png", b"fake image data")

        result = self.parser.parse_file(zip_file)

        self.assertIsNotNone(result)
        self.assertEqual(result["extension"], ".zip")
        self.assertIn("ZIP Archive: test_archive.zip", result["content"])
        self.assertIn("readme.md", result["content"])
        self.assertIn("notes.txt", result["content"])
        self.assertIn("config.json", result["content"])
        self.assertIn("Test Documentation", result["content"])
        self.assertIn("Important notes here", result["content"])
        self.assertIn("image.png (unsupported format)", result["content"])

    def test_parse_tar_archive(self):
        """Test parsing TAR archives containing supported documents."""
        # Create a TAR file with test documents
        tar_file = self.temp_path / "test_archive.tar"

        with tarfile.open(tar_file, "w") as tar_ref:
            # Create temporary files to add to the archive
            md_content = "# TAR Test\n\nThis is from a TAR file."
            txt_content = "Text file in TAR archive."

            # Add files using tarfile's string support
            md_info = tarfile.TarInfo(name="documentation.md")
            md_info.size = len(md_content.encode())
            tar_ref.addfile(md_info, fileobj=self._string_to_fileobj(md_content))

            txt_info = tarfile.TarInfo(name="notes.txt")
            txt_info.size = len(txt_content.encode())
            tar_ref.addfile(txt_info, fileobj=self._string_to_fileobj(txt_content))

        result = self.parser.parse_file(tar_file)

        self.assertIsNotNone(result)
        self.assertEqual(result["extension"], ".tar")
        self.assertIn("TAR Archive: test_archive.tar", result["content"])
        self.assertIn("documentation.md", result["content"])
        self.assertIn("notes.txt", result["content"])
        self.assertIn("TAR Test", result["content"])
        self.assertIn("Text file in TAR archive", result["content"])

    def test_parse_tar_gz_archive(self):
        """Test parsing compressed TAR archives."""
        # Create a TAR.GZ file with test documents
        tar_file = self.temp_path / "test_archive.tar.gz"

        with tarfile.open(tar_file, "w:gz") as tar_ref:
            csv_content = "name,value\ntest,123\nexample,456"

            csv_info = tarfile.TarInfo(name="data.csv")
            csv_info.size = len(csv_content.encode())
            tar_ref.addfile(csv_info, fileobj=self._string_to_fileobj(csv_content))

        result = self.parser.parse_file(tar_file)

        self.assertIsNotNone(result)
        self.assertEqual(result["extension"], ".tar.gz")
        self.assertIn("TAR Archive: test_archive.tar.gz", result["content"])
        self.assertIn("data.csv", result["content"])
        self.assertIn("name, value", result["content"])

    def test_empty_zip_archive(self):
        """Test parsing empty ZIP archives."""
        zip_file = self.temp_path / "empty.zip"

        with zipfile.ZipFile(zip_file, "w") as zip_ref:
            pass  # Create empty ZIP

        result = self.parser.parse_file(zip_file)

        self.assertIsNotNone(result)
        self.assertEqual(result["extension"], ".zip")
        self.assertIn("Contains 0 files", result["content"])

    def test_zip_with_nested_directories(self):
        """Test ZIP archives with nested directory structures."""
        zip_file = self.temp_path / "nested.zip"

        with zipfile.ZipFile(zip_file, "w") as zip_ref:
            zip_ref.writestr("project/docs/api.md", "# API Documentation")
            zip_ref.writestr("project/src/config.json", '{"debug": true}')
            zip_ref.writestr("project/README.txt", "Project readme file")

        result = self.parser.parse_file(zip_file)

        self.assertIsNotNone(result)
        self.assertIn("project/docs/api.md", result["content"])
        self.assertIn("project/src/config.json", result["content"])
        self.assertIn("project/README.txt", result["content"])
        self.assertIn("API Documentation", result["content"])

    def _string_to_fileobj(self, content: str):
        """Helper method to convert string to file-like object for tarfile."""
        import io

        return io.BytesIO(content.encode("utf-8"))

    def test_parse_html(self):
        """Test parsing HTML files."""
        content = """
        <!DOCTYPE html>
        <html>
        <head><title>Test Page</title></head>
        <body>
            <h1>Welcome</h1>
            <p>This is a test HTML page.</p>
            <div>Some content in a div.</div>
        </body>
        </html>
        """
        html_file = self.temp_path / "test.html"
        html_file.write_text(content)

        result = self.parser.parse_file(html_file)

        self.assertIsNotNone(result)
        self.assertIn("Welcome", result["content"])
        self.assertIn("test HTML page", result["content"])
        self.assertEqual(result["extension"], ".html")

    def test_parse_xml(self):
        """Test parsing XML files."""
        content = """<?xml version="1.0" encoding="UTF-8"?>
        <project>
            <name>Test Project</name>
            <version>1.0.0</version>
            <description>A test project configuration</description>
            <dependencies>
                <dependency>
                    <name>library1</name>
                    <version>2.0.0</version>
                </dependency>
            </dependencies>
        </project>"""
        xml_file = self.temp_path / "test.xml"
        xml_file.write_text(content)

        result = self.parser.parse_file(xml_file)

        self.assertIsNotNone(result)
        self.assertIn("Test Project", result["content"])
        self.assertIn("library1", result["content"])
        self.assertEqual(result["extension"], ".xml")

    def test_parse_xlsx(self):
        """Test parsing XLSX files."""
        try:
            import openpyxl

            # Create a simple XLSX file
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test Sheet"

            # Add some test data
            ws["A1"] = "Name"
            ws["B1"] = "Age"
            ws["C1"] = "Department"
            ws["A2"] = "John Doe"
            ws["B2"] = 30
            ws["C2"] = "Engineering"
            ws["A3"] = "Jane Smith"
            ws["B3"] = 25
            ws["C3"] = "Marketing"

            xlsx_file = self.temp_path / "test.xlsx"
            wb.save(xlsx_file)

            result = self.parser.parse_file(xlsx_file)

            self.assertIsNotNone(result)
            self.assertIn("Test Sheet", result["content"])
            self.assertIn("John Doe", result["content"])
            self.assertIn("Engineering", result["content"])
            self.assertEqual(result["extension"], ".xlsx")

        except ImportError:
            self.skipTest("openpyxl not available for XLSX testing")


if __name__ == "__main__":
    unittest.main()
